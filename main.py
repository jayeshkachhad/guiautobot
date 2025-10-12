import pyautogui
import time
import webbrowser
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Get the current date and time

wb = load_workbook("zomatoNums.xlsx")
ws = wb.active

screenWidth, screenHeight = pyautogui.size()
print(f"Screen size: {screenWidth}x{screenHeight}")
start = datetime.now()
print(start)

# message = "Are you a Shopify pro who can turn ideas into pixel-perfect, high-converting eCommerce stores? 🧑‍💻🛒 \n We’re currently looking for \n 🎯 Position: Shopify Developer (1+ Years Experience) \n Location: Surat (Onsite) \n 📩 Apply Now: hr@novicbase.com or WhatsApp: What We’re Looking For: \n ✅ Hands-on experience with Shopify & Shopify Plus \n ✅ Custom theme development & Liquid expertise \n ✅ Knowledge of HTML, CSS, JS, and APIs \n ✅ Ability to troubleshoot, optimize & scale eCommerce solutions \n ✅ Bonus: Shopify app integration and custom app experience \n If you’re someone who’s passionate about clean code, user experience, and eCommerce innovation — we’d love to hear from you! 🚀 \n Join us at Novicbase and help shape the future of online🌐"

for i in range(ws.max_row):
    name = (ws.cell(row=i+1, column=1)
            .value)
    phone = ws.cell(row=i+1, column=2).value
    status = ws.cell(row=i+1, column=3).value

    print("Sr. ", i,  name)
    if(status == 'Done'):
        continue
    else:
        # time.sleep(4)
        # ss = pyautogui.screenshot()
        # findbt = pyautogui.locateOnScreen('NOTfoundBig.png', confidence=0.9)
        print("Performing", i, name)
        try:
            button_location = pyautogui.locateOnScreen('ok.png', confidence=0.9, grayscale=False)
            if button_location:
                time.sleep(3)
                print(f"Button found at: {button_location}")
                # You can then click its center:
                try:
                    pyautogui.click(button_location)
                    print("Clicked the button!")
                except:
                    continue
                time.sleep(2)
            else:
                print("Button not found on screen.")
        except pyautogui.ImageNotFoundException:
            print("Button not found (using default exception handling).")
        time.sleep(2)
        url = 'whatsapp://send?phone=' + str(phone)
        time.sleep(2)
        webbrowser.open(url)
        time.sleep(2)

        try:
            mf = pyautogui.locateOnScreen('messageField.png', confidence=0.9, grayscale=False)
            if mf:
                # pyautogui.write()3
                pyautogui.moveTo(mf.top, mf.left, duration=0.5)

                text = [
                    "Let AI Generate Your Digital Food Menu!",
                    "Convert Your Old menu to fully Digital Menu in just minutes.",
                    "Send Your menu PDF Or Images and we will take care of rest."
                ]
                # pyautogui.typewrite(" Convert Your Old menu to fully digital menu in just 2 minutes. Send Your menu PDF Or Images and we will take care of rest. ")
                pyautogui.hotkey('ctrl', 'v')
                # for line in text:
                #     pyautogui.typewrite(line)
                #     pyautogui.hotkey('shift', 'enter')  # Insert newline without sending
                #     time.sleep(2.5)

                time.sleep(3)
                pyautogui.hotkey('enter')
                ws.cell(row=i+1, column=3).value = "Done"
                ws.cell(row=i+1, column=4).value = datetime.now()
                wb.save("zomatoNums.xlsx")
                time.sleep(3)
            else:
                print("Write Field Not Available")
                time.sleep(4)
        except:
            time.sleep(3)
            continue

print("End @ ", datetime.now())
# os.system("shutdown /s /t 0")